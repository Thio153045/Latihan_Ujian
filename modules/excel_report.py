"""Membuat rekap nilai dalam bentuk Excel (.xlsx) untuk guru — dipakai saat
menganalisis/melihat nilai LEBIH DARI SATU siswa sekaligus, supaya tidak
perlu buka PDF satu-satu untuk sekadar melihat rekap skor."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def generate_recap_excel(rows: list) -> io.BytesIO:
    """rows: list of dict dengan keys nama_lengkap, kelas, nama_sekolah,
    paket_nama, skor, total_soal, dianalisis_at (atau dikerjakan_at)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Nilai"

    headers = ["No", "Nama Siswa", "Kelas", "Sekolah", "Paket Soal",
               "Skor", "Total Soal", "Persentase", "Tanggal Dinilai"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=1):
        skor = r.get("skor")
        total = r.get("total_soal") or 0
        pct = round(100 * skor / total, 1) if (skor is not None and total) else None
        tanggal = r.get("dianalisis_at") or r.get("dikerjakan_at") or ""
        ws.append([
            i, r.get("nama_lengkap", ""), r.get("kelas", ""), r.get("nama_sekolah", ""),
            r.get("paket_nama", ""), skor, total,
            f"{pct}%" if pct is not None else "-", str(tanggal),
        ])

    widths = [5, 22, 10, 22, 26, 8, 10, 11, 20]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
